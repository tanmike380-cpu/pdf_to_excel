"""Excel export service."""

import os
import re
from pathlib import Path
from typing import List, Dict

import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

from app.schemas.response import StandardParseResult
from app.utils.temp_files import get_excel_path, ensure_dirs
from app.core.logger import get_logger

logger = get_logger("excel_service")


def sanitize_sheet_title(title: str, max_length: int = 31) -> str:
    """Sanitize sheet title: remove illegal chars, limit length."""
    # Excel sheet name rules: no \ / ? * [ ]
    illegal = r'[\\\/?*\[\]:]'
    title = re.sub(illegal, '_', title)
    title = title.strip()
    if not title:
        title = "Sheet1"
    return title[:max_length]


def export_to_excel(
    result: StandardParseResult,
    file_id: str,
    sheet_title: str = "Extraction Result",
) -> Path:
    """
    Export StandardParseResult to Excel file.
    
    Returns path to the generated Excel file.
    """
    ensure_dirs()
    
    sheet_name = sanitize_sheet_title(sheet_title)
    excel_path = get_excel_path(file_id)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    
    # Write headers
    for col_idx, col_name in enumerate(result.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Write data rows
    for row_idx, row_data in enumerate(result.rows, start=2):
        for col_idx, col_name in enumerate(result.columns, start=1):
            value = row_data.get(col_name, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    # Auto-adjust column widths
    for col_idx, col_name in enumerate(result.columns, start=1):
        max_length = len(col_name)
        for row in result.rows:
            cell_value = str(row.get(col_name, ""))
            # Consider line breaks
            for line in cell_value.split('\n'):
                max_length = max(max_length, len(line))
        
        # Cap at 50 chars width
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width
    
    # Freeze first row
    ws.freeze_panes = 'A2'
    
    wb.save(excel_path)
    logger.info(f"Excel saved: {excel_path}")
    
    return excel_path


def get_preview_rows(result: StandardParseResult, max_rows: int = 20) -> List[Dict[str, str]]:
    """Get first N rows for preview."""
    return result.rows[:max_rows]
